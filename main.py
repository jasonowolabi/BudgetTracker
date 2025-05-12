from openpyxl import *
import time

date = time.strftime("%m/%d")
currTime = time.strftime("%I:%M %p")

wb = load_workbook('db.xlsx')
ws = wb.active

balance = ws['B2'].value
savings = ws['B3'].value 

def update(item, price):
    ws.append(date, item, price)
    ws['B2'].value = balance - price

def savings(price):
    ws['B3'].value = savings + price 
    

print(balance, savings)

wb.save('db.xlsx')

